"""Excel report generator for ePetCare desktop application.

This module provides utilities for generating Excel reports of patients and owners.
"""

import os
import logging
from datetime import datetime
from typing import Tuple
from pathlib import Path

logger = logging.getLogger('epetcare')

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available. Excel report generation will be disabled.")


def get_reports_dir() -> str:
    """Return the reports directory path, creating it if it doesn't exist."""
    reports_dir = Path(os.path.dirname(os.path.dirname(__file__))) / "reports"
    reports_dir.mkdir(exist_ok=True)
    return str(reports_dir)


def generate_patients_report_to_path(file_path: str) -> Tuple[bool, str]:
    """Generate an Excel report of all patients and their owners to a specific path.
    
    Args:
        file_path: The full path where the Excel file should be saved.
    
    Returns:
        Tuple of (success, result). If success is True, result contains the
        path to the Excel file. If success is False, result contains an error message.
    """
    if not OPENPYXL_AVAILABLE:
        return False, "openpyxl library not installed. Please install it with: pip install openpyxl"
    
    try:
        # Initialize PostgreSQL connection
        from utils.config import load_config
        from utils.pg_db import setup_postgres_connection
        
        config = load_config()
        pg_config = config.get('postgres', {})
        
        if not pg_config:
            return False, "PostgreSQL configuration not found in config.json"
        
        # Setup connection
        if not setup_postgres_connection(pg_config):
            return False, "Failed to connect to PostgreSQL database"
        
        from models.data_access import PetDataAccess, OwnerDataAccess
        
        # Initialize data access
        pet_data = PetDataAccess()
        owner_data = OwnerDataAccess()
        
        # Fetch all patients with owner information
        logger.info("Fetching all patients for report...")
        from models.models import Pet, Owner
        
        # Fetch all pets directly from database
        success, result = pet_data.db.fetch_all('clinic_pet')
        
        if not success:
            return False, f"Failed to fetch patients: {result}"
        
        # Convert rows to Pet objects
        patients = []
        for row in result:
            pet_dict = pet_data._row_to_dict(row)
            patients.append(Pet(
                id=pet_dict['id'],
                owner_id=pet_dict['owner_id'],
                name=pet_dict['name'],
                species=pet_dict['species'],
                breed=pet_dict['breed'],
                sex=pet_dict['sex'],
                birth_date=pet_data._parse_date(pet_dict['birth_date']) if pet_dict['birth_date'] else None,
                weight_kg=float(pet_dict['weight_kg']) if pet_dict['weight_kg'] else None,
                notes=pet_dict['notes'],
                image=pet_dict.get('image', None)
            ))
        
        if not patients:
            return False, "No patients found in the database."
        
        # Fetch owner information directly from database
        success, owner_rows = owner_data.db.fetch_all('clinic_owner')
        owners_dict = {}
        
        if success and owner_rows:
            for row in owner_rows:
                owner_dict = owner_data._row_to_dict(row)
                owners_dict[owner_dict['id']] = Owner(
                    id=owner_dict['id'],
                    full_name=owner_dict.get('full_name', ''),
                    email=owner_dict['email'],
                    phone=owner_dict['phone'],
                    address=owner_dict.get('address', ''),
                    created_at=owner_data._parse_datetime(owner_dict.get('created_at')),
                    user_id=owner_dict.get('user_id')
                )
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Create Patients sheet
        ws_patients = wb.active
        ws_patients.title = "Patients"
        
        # Define headers for patients
        patient_headers = [
            "ID", "Name", "Species", "Breed", "Sex", "Birth Date", 
            "Weight (kg)", "Notes", "Owner ID", "Owner Name", "Owner Contact"
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write patient headers
        for col_num, header in enumerate(patient_headers, 1):
            cell = ws_patients.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write patient data
        for row_num, patient in enumerate(patients, 2):
            owner = owners_dict.get(patient.owner_id) if patient.owner_id else None
            
            ws_patients.cell(row=row_num, column=1).value = patient.id
            ws_patients.cell(row=row_num, column=2).value = patient.name
            ws_patients.cell(row=row_num, column=3).value = patient.species
            ws_patients.cell(row=row_num, column=4).value = patient.breed
            ws_patients.cell(row=row_num, column=5).value = patient.sex
            ws_patients.cell(row=row_num, column=6).value = str(patient.birth_date) if patient.birth_date else ""
            ws_patients.cell(row=row_num, column=7).value = float(patient.weight_kg) if patient.weight_kg else ""
            ws_patients.cell(row=row_num, column=8).value = patient.notes or ""
            ws_patients.cell(row=row_num, column=9).value = patient.owner_id
            ws_patients.cell(row=row_num, column=10).value = owner.full_name if owner else ""
            ws_patients.cell(row=row_num, column=11).value = owner.phone if owner else ""
        
        # Auto-adjust column widths
        for col_num in range(1, len(patient_headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws_patients[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_patients.column_dimensions[column_letter].width = adjusted_width
        
        # Create Owners sheet
        ws_owners = wb.create_sheet(title="Owners")
        
        # Define headers for owners
        owner_headers = [
            "ID", "First Name", "Last Name", "Email", "Phone", 
            "Address", "Number of Pets"
        ]
        
        # Write owner headers
        for col_num, header in enumerate(owner_headers, 1):
            cell = ws_owners.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Count pets per owner
        pets_per_owner = {}
        for patient in patients:
            if patient.owner_id:
                pets_per_owner[patient.owner_id] = pets_per_owner.get(patient.owner_id, 0) + 1
        
        # Write owner data
        for row_num, owner in enumerate(owners_dict.values(), 2):
            # Split full name for display
            name_parts = owner.full_name.split(' ', 1) if owner.full_name else ['', '']
            first_name = name_parts[0] if name_parts else ''
            last_name = name_parts[1] if len(name_parts) > 1 else ''
            
            ws_owners.cell(row=row_num, column=1).value = owner.id
            ws_owners.cell(row=row_num, column=2).value = first_name
            ws_owners.cell(row=row_num, column=3).value = last_name
            ws_owners.cell(row=row_num, column=4).value = owner.email
            ws_owners.cell(row=row_num, column=5).value = owner.phone
            ws_owners.cell(row=row_num, column=6).value = owner.address or ""
            ws_owners.cell(row=row_num, column=7).value = pets_per_owner.get(owner.id, 0)
        
        # Auto-adjust column widths for owners sheet
        for col_num in range(1, len(owner_headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws_owners[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_owners.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook to specified path
        wb.save(file_path)
        
        logger.info(f"Report generated successfully: {file_path}")
        logger.info(f"Report contains {len(patients)} patients and {len(owners_dict)} owners")
        
        return True, file_path
        
    except Exception as e:
        logger.error(f"Failed to generate report: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False, f"Failed to generate report: {str(e)}"


def generate_patients_report() -> Tuple[bool, str]:
    """Generate an Excel report of all patients and their owners.
    
    Returns:
        Tuple of (success, result). If success is True, result contains the
        path to the Excel file. If success is False, result contains an error message.
    """
    if not OPENPYXL_AVAILABLE:
        return False, "openpyxl library not installed. Please install it with: pip install openpyxl"
    
    try:
        # Initialize PostgreSQL connection
        from utils.config import load_config
        from utils.pg_db import setup_postgres_connection
        
        config = load_config()
        pg_config = config.get('postgres', {})
        
        if not pg_config:
            return False, "PostgreSQL configuration not found in config.json"
        
        # Setup connection
        if not setup_postgres_connection(pg_config):
            return False, "Failed to connect to PostgreSQL database"
        
        from models.data_access import PetDataAccess, OwnerDataAccess
        
        # Initialize data access
        pet_data = PetDataAccess()
        owner_data = OwnerDataAccess()
        
        # Fetch all patients with owner information
        logger.info("Fetching all patients for report...")
        from models.models import Pet, Owner
        
        # Fetch all pets directly from database
        success, result = pet_data.db.fetch_all('clinic_pet')
        
        if not success:
            return False, f"Failed to fetch patients: {result}"
        
        # Convert rows to Pet objects
        patients = []
        for row in result:
            pet_dict = pet_data._row_to_dict(row)
            patients.append(Pet(
                id=pet_dict['id'],
                owner_id=pet_dict['owner_id'],
                name=pet_dict['name'],
                species=pet_dict['species'],
                breed=pet_dict['breed'],
                sex=pet_dict['sex'],
                birth_date=pet_data._parse_date(pet_dict['birth_date']) if pet_dict['birth_date'] else None,
                weight_kg=float(pet_dict['weight_kg']) if pet_dict['weight_kg'] else None,
                notes=pet_dict['notes'],
                image=pet_dict.get('image', None)
            ))
        
        if not patients:
            return False, "No patients found in the database."
        
        # Fetch owner information directly from database
        success, owner_rows = owner_data.db.fetch_all('clinic_owner')
        owners_dict = {}
        
        if success and owner_rows:
            for row in owner_rows:
                owner_dict = owner_data._row_to_dict(row)
                owners_dict[owner_dict['id']] = Owner(
                    id=owner_dict['id'],
                    full_name=owner_dict.get('full_name', ''),
                    email=owner_dict['email'],
                    phone=owner_dict['phone'],
                    address=owner_dict.get('address', ''),
                    created_at=owner_data._parse_datetime(owner_dict.get('created_at')),
                    user_id=owner_dict.get('user_id')
                )
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Create Patients sheet
        ws_patients = wb.active
        ws_patients.title = "Patients"
        
        # Define headers for patients
        patient_headers = [
            "ID", "Name", "Species", "Breed", "Sex", "Birth Date", 
            "Weight (kg)", "Notes", "Owner ID", "Owner Name", "Owner Contact"
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write patient headers
        for col_num, header in enumerate(patient_headers, 1):
            cell = ws_patients.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write patient data
        for row_num, patient in enumerate(patients, 2):
            owner = owners_dict.get(patient.owner_id) if patient.owner_id else None
            
            ws_patients.cell(row=row_num, column=1).value = patient.id
            ws_patients.cell(row=row_num, column=2).value = patient.name
            ws_patients.cell(row=row_num, column=3).value = patient.species
            ws_patients.cell(row=row_num, column=4).value = patient.breed
            ws_patients.cell(row=row_num, column=5).value = patient.sex
            ws_patients.cell(row=row_num, column=6).value = str(patient.birth_date) if patient.birth_date else ""
            ws_patients.cell(row=row_num, column=7).value = float(patient.weight_kg) if patient.weight_kg else ""
            ws_patients.cell(row=row_num, column=8).value = patient.notes or ""
            ws_patients.cell(row=row_num, column=9).value = patient.owner_id
            ws_patients.cell(row=row_num, column=10).value = owner.full_name if owner else ""
            ws_patients.cell(row=row_num, column=11).value = owner.phone if owner else ""
        
        # Auto-adjust column widths
        for col_num in range(1, len(patient_headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws_patients[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_patients.column_dimensions[column_letter].width = adjusted_width
        
        # Create Owners sheet
        ws_owners = wb.create_sheet(title="Owners")
        
        # Define headers for owners
        owner_headers = [
            "ID", "First Name", "Last Name", "Email", "Phone", 
            "Address", "Number of Pets"
        ]
        
        # Write owner headers
        for col_num, header in enumerate(owner_headers, 1):
            cell = ws_owners.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Count pets per owner
        pets_per_owner = {}
        for patient in patients:
            if patient.owner_id:
                pets_per_owner[patient.owner_id] = pets_per_owner.get(patient.owner_id, 0) + 1
        
        # Write owner data
        for row_num, owner in enumerate(owners_dict.values(), 2):
            # Split full name for display
            name_parts = owner.full_name.split(' ', 1) if owner.full_name else ['', '']
            first_name = name_parts[0] if name_parts else ''
            last_name = name_parts[1] if len(name_parts) > 1 else ''
            
            ws_owners.cell(row=row_num, column=1).value = owner.id
            ws_owners.cell(row=row_num, column=2).value = first_name
            ws_owners.cell(row=row_num, column=3).value = last_name
            ws_owners.cell(row=row_num, column=4).value = owner.email
            ws_owners.cell(row=row_num, column=5).value = owner.phone
            ws_owners.cell(row=row_num, column=6).value = owner.address or ""
            ws_owners.cell(row=row_num, column=7).value = pets_per_owner.get(owner.id, 0)
        
        # Auto-adjust column widths for owners sheet
        for col_num in range(1, len(owner_headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws_owners[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_owners.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook to default reports directory
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        reports_dir = get_reports_dir()
        report_filename = f"patients_report_{timestamp}.xlsx"
        report_path = os.path.join(reports_dir, report_filename)
        
        # Use the new function to save to the default path
        return generate_patients_report_to_path(report_path)
        
    except Exception as e:
        logger.error(f"Failed to generate report: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False, f"Failed to generate report: {str(e)}"
